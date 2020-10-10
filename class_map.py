# -*- coding: UTF-8 -*-
#import PyPDF2
import time
import datetime
#from zipfile import ZipFile
import shutil
import xlrd
#from PIL import Image
import openpyxl
from pathlib import Path
import re
import os
from io import FileIO
import docx
from docx.enum.text import WD_ALIGN_PARAGRAPH
from docx.shared import Inches
import pandas
from docx import Document
from tkinter import filedialog
from tkinter import messagebox
#from aip import AipOcr
import matplotlib.pyplot as plt
import matplotlib
from pyecharts.charts import Map,Geo,Timeline
from pyecharts import options as opts


class Import_data():
	def __init__(self,file):
		self.file          =file

	def map_data(self):
		data_file          =self.file
		df=pandas.read_excel(data_file)
		wb=openpyxl.load_workbook(data_file)
		sheet1=wb[wb.sheetnames[0]]
		A1=sheet1['A1'].value
		B1=sheet1['B1'].value
		part=list(df[A1])
		element=list(df[B1])
		r=0
		#for el in element:
		#	r+=int(el)

		r=int(max(element))
		prov = ['河北',
			'山西',
			'辽宁',
			'吉林',
			'黑龙江',
			'江苏',
			'浙江',
			'安徽',
			'福建',
			'江西',
			'山东',
			'河南',
			'湖北',
			'湖南',
			'广东',
			'海南',
			'四川',
			'贵州',
			'云南',
			'陕西',
			'甘肃',
			'青海',
			'台湾',
			'内蒙古',
			'广西',
			'西藏',
			'宁夏',
			'新疆',
			'北京',
			'天津',
			'上海',
			'重庆',
			'香港',
			'澳门',
			'台湾',
			'南海诸岛']
		part_anal=[]
		if len(part) == len(element):
			for j,t in enumerate(part):
				for p in prov:
					if t in p:
						tup=(t,int(element[j]))
						part_anal.append(tup)
						tup=()
					else:
						tup=(p,0)
						part_anal.append(tup)
						tup=()
		#(part_anal)
		#print(r)
		#r=int(r)
		return part_anal,r

	def color_data(self):
		data_file          =self.file
		df=pandas.read_excel(data_file)
		wb=openpyxl.load_workbook(data_file)
		sheet1=wb[wb.sheetnames[0]]
		A1=sheet1['A1'].value
		B1=sheet1['B1'].value
		part=list(df[A1])
		element=list(df[B1])
		r=0
		for el in element:
			r+=int(el)

		
		return r

		
		

class Data_map():
	def __init__(self,dirs):
		self.dirs                      =dirs

	def result(self):
		dirs                           =self.dirs
		excel_list=[f for f in Path(dirs).glob('**/*.xlsx')]
		for di in excel_list:
			dirname,filename=os.path.split(di)
			filename_1=filename.split('.')
			if filename_1[0]!='':
				my_data    =Import_data(di)
				part_anal,r=my_data.map_data()
				return part_anal,r

class Map_html():
	def __init__(self,dirs,rname,data,max_l):
		self.dirs                      =dirs
		self.rname                     =rname
		self.data                      =data
		self.max_l                     =max_l

	def china(self):
		dirs                           =self.dirs
		rname                          =self.rname
		data                           =self.data
		max_l                          =self.max_l
		rich={
			"{b}":{"color": "red"}
		}
		os.chdir(dirs)
		itemstyle_opts={
			"normal": {"areaColor": "#323c48","borderColor": "#404a59"},
			"emphasis": {
				"label": {"show": Timeline},
				"areaColor": "rgba(255,255,255, 0.5)"
				}
			}
		map=Map(init_opts=opts.InitOpts(width="900px",height="800px"))
		map.add("",data,maptype='china',is_map_symbol_show=False,label_opts= opts.LabelOpts(is_show=True,formatter="{b}\n{c}",rich=rich, font_size=10))
		#map.Item(name,value)
		map.set_global_opts(
    		title_opts=opts.TitleOpts(title="",subtitle="",pos_right="center",pos_top="5%"),
    		visualmap_opts=opts.VisualMapOpts(max_=max_l,range_color=["white","skyblue"],is_piecewise=False),
    		)
		map.render(rname+".html")

def main():
	#dirs      =filedialog.askdirectory()
	dirs      ='C:\\Users\\woshi\\Desktop\\地图'
	data      =[]
	max_l=0
	my_m      =Data_map(dirs)
	data,max_l=my_m.result()
	print(data)
	#data=[("浙江",123),("贵州"，145)]
	rname     ='地图绘制结果'
	my_map    =Map_html(dirs,rname,data,max_l)
	my_map.china()

if __name__ == "__main__":
	main()



